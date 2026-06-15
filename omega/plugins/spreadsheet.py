from __future__ import annotations

import csv
from typing import Any

from omega.plugins.base import Plugin, PluginContext
from omega.schema import PluginResult


class SpreadsheetPlugin(Plugin):
    name = "spreadsheet"
    description = "Read and write CSV files and read XLSX files when openpyxl is installed."
    actions = {
        "read_csv": "Read a CSV file.",
        "write_csv": "Write a CSV file.",
        "read_xlsx": "Read an XLSX worksheet.",
    }

    async def call(self, action: str, arguments: dict[str, Any], context: PluginContext) -> PluginResult:
        if action == "read_csv":
            path = context.permissions.resolve_path(arguments["path"])
            with path.open(newline="", encoding=arguments.get("encoding", "utf-8")) as handle:
                rows = list(csv.DictReader(handle)) if arguments.get("headers", True) else list(csv.reader(handle))
            return PluginResult(plugin=self.name, action=action, ok=True, data=rows)
        if action == "write_csv":
            path = context.permissions.resolve_path(arguments["path"])
            path.parent.mkdir(parents=True, exist_ok=True)
            rows = list(arguments.get("rows", []))
            with path.open("w", newline="", encoding=arguments.get("encoding", "utf-8")) as handle:
                if rows and isinstance(rows[0], dict):
                    writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                else:
                    csv.writer(handle).writerows(rows)
            return PluginResult(plugin=self.name, action=action, ok=True, data={"path": str(path), "rows": len(rows)})
        if action == "read_xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                return PluginResult(plugin=self.name, action=action, ok=False, error="openpyxl is not installed.")
            path = context.permissions.resolve_path(arguments["path"])
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook[arguments.get("sheet")] if arguments.get("sheet") else workbook.active
            rows = [[cell for cell in row] for row in worksheet.iter_rows(values_only=True)]
            return PluginResult(plugin=self.name, action=action, ok=True, data={"sheet": worksheet.title, "rows": rows})
        return self.unknown_action(action)
